#!/usr/bin/env python3
"""
Rigenera i coefficienti BCC Rent&Lease in src/data/bcc.ts a partire dal planner.

Quando serve: ogni volta che BCC pubblica un planner nuovo (nuovo trimestre o
nuova campagna). Il planner non contiene i coefficienti come tabella: li calcola
col modello (provvista -> TIR -> cash flow). Questo script ricalcola il modello
su ogni combinazione di prodotto, classe di rischio, durata e fascia di importo.

Uso:
    pip3 install --user formulas openpyxl
    python3 scripts/genera-coefficienti-bcc.py "/percorso/del/planner.xlsx"

Senza argomenti usa il planner di default in ~/Desktop/AREE/Nord_Est_Group/BCC/.

Verifica automatica: ricalcola il caso noto (75.000 EUR a 60 mesi classe 3 in
locazione operativa = 1.540,06) e si ferma se non torna.
"""
import sys, os, json, tempfile, warnings
warnings.filterwarnings('ignore')

import openpyxl
import formulas
from formulas.parser import Parser

DEFAULT = os.path.expanduser(
    '~/Desktop/AREE/Nord_Est_Group/BCC/#_BCCRLplanner_LO+LF+FF_260226.xlsx')
DEST = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'bcc.ts')

# Config di calcolo: la stessa del preventivatore usato in operativo
BASE = dict(B24='Mensile', B26='primo canone ', B32=2, B33='SI', B34=0.03, B35=0)
PRODOTTI = {
    'lo': ('LOCAZIONE OPERATIVA', 0.01),
    'lf': ('LOCAZIONE FINANZIARIA', 0.01),
    'ff': ('FINANZIAMENTO FINALIZZATO', 0.0),
}
FASCE = [('<= 5.000', 4000, 3000), ('5.001 - 15.000', 10000, 13000),
         ('15.001 - 25.000', 20000, 23000), ('25.001 - 50.000', 35000, 45000),
         ('50.001 - 100.000', 75000, 90000), ('> 100.000', 150000, 180000)]
DURATE = [18, 24, 30, 36, 48, 60]
CLASSI = [1, 2, 3, 4, 5]


def sanifica(src):
    """Congela le formule che il parser non digerisce (sono solo di visualizzazione)."""
    wf = openpyxl.load_workbook(src, data_only=False)
    wv = openpyxl.load_workbook(src, data_only=True)
    p, n = Parser(), 0
    for ws in wf.worksheets:
        wsv = wv[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    try:
                        p.ast(cell.value)
                    except Exception:
                        cell.value = wsv[cell.coordinate].value
                        n += 1
    dst = os.path.join(tempfile.mkdtemp(), 'planner.xlsx')
    wf.save(dst)
    print(f'formule congelate: {n}')
    return dst


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT
    print('planner:', src)
    path = sanifica(src)
    xl = formulas.ExcelModel().loads(path).finish()
    F = '[planner.xlsx]'
    K = lambda sh, c: f"'{F}{sh}'!{c}"
    OUT = [K('CASH FLOW', 'B43')]

    def rata(nome, imp, dur, vr, classe):
        ins = dict(B18=nome, B19=classe, B23=imp, B25=dur, B28=vr, **BASE)
        sol = xl.calculate(inputs={K('INPUT', c): v for c, v in ins.items()}, outputs=OUT)
        try:
            return float(sol[OUT[0]].value[0, 0])
        except Exception:
            return float('nan')

    atteso = rata('LOCAZIONE OPERATIVA', 75000, 60, 0.01, 3)
    if abs(atteso - 1540.06) > 0.02:
        sys.exit(f'VERIFICA FALLITA: 75.000 a 60 mesi classe 3 da {atteso:.2f}, atteso 1540,06. '
                 'Il planner e cambiato in modo strutturale: controllare prima di rigenerare.')
    print(f'verifica ok: caso noto = {atteso:.2f}')

    coeff = {}
    for pk, (nome, vr) in PRODOTTI.items():
        coeff[pk] = {}
        for cl in CLASSI:
            coeff[pk][cl] = {}
            for d in DURATE:
                riga = []
                for _, imp, imp2 in FASCE:
                    c = rata(nome, imp, d, vr, cl)
                    v = round(c / imp * 100, 4) if c == c and c > 0 else None
                    c2 = rata(nome, imp2, d, vr, cl)
                    v2 = round(c2 / imp2 * 100, 4) if c2 == c2 and c2 > 0 else None
                    if v and v2 and abs(v - v2) > 0.004:
                        print(f'  ATTENZIONE coefficiente non costante in fascia: {pk} cl{cl} {d}m {v} vs {v2}')
                    riga.append(v)
                coeff[pk][cl][d] = riga
        print(f'{pk} completato')

    print('\nCoefficienti pronti. Aggiornare a mano le griglie in', os.path.normpath(DEST))
    json.dump(coeff, open('coefficienti_bcc.json', 'w'), indent=1)
    print('dump JSON: coefficienti_bcc.json')


if __name__ == '__main__':
    main()
